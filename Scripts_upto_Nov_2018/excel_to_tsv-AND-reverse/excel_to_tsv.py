'''
The script is written for extracting the required data from 
.XLSX files and create .TSV files for the data.
'''
import openpyxl
from openpyxl import load_workbook
import csv
import re
import glob

xl_files = sorted(glob.glob("*.xlsx"))

for xl_file in xl_files:
	work_book = load_workbook(xl_file)
	sheets = work_book.get_sheet_names()
	work_sheet = work_book.get_sheet_by_name(sheets[0])
	work_book.active
	file_name = re.sub(r"\.xlsx","",xl_file)
	
	'''Creating a csv file'''
	with open (file_name + '.tsv', 'w', encoding = 'utf-8') as tsv_file:
		twriter = csv.writer(tsv_file, delimiter='\t')

		for row in range(1,work_sheet.max_row + 1):
			book = work_sheet['A' + str(row)].value
			chapter = work_sheet['B' + str(row)].value
			verse = work_sheet['C' + str(row)].value
			note = work_sheet['D' + str(row)].value

			'''Using Regex for cleaning-up the unwanted contents'''
			edited_note = re.sub(r"\$","<br>",note)

			print(book,chapter,verse,edited_note)
			twriter.writerow([book,chapter,verse,edited_note]);
