'''
The script is for converting CSV file's in BCV format to USFM files 
by adding the required tags and contents like book ID and book full name in the
same language for multiple bible's.
'''
import csv
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import glob
import os
import re

'''Function is for searching the language of the file'''
def search_lang(lang):
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for column in "CDEFGHIJKLMNOPQ":
			cell_name = "{}{}".format(column, 1)
			language = worksheet[cell_name].value.split("/")

			''' It compare(s) the language with the .xlsx file '''
			if language[1].lower() == lang.lower():
				print(language[1],lang)
				return column

''' Function is for updating the book name in same language '''
def search_name(book,col):
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for row in range(2,worksheet.max_row + 1):	
			book_num = worksheet["A" + str(row)].value
			
			''' Compares the book name with excel sheet 
			    and returns the book name in same languag '''
			if str(book) == str(book_num):
				book_name = worksheet[col + str(row)].value
				book_id = worksheet["B" + str(row)].value
				#os.chdir("docx")
				#os.chdir(path)
				return book_name,book_id


csv_files = glob.glob("*.csv")

for csv_file in csv_files:
	extract_lang = re.match(r"(\w+_)(\w+)(_\w+)(\.csv)",csv_file)
	if extract_lang:
		csv_lang = extract_lang.group(2)

	''' Search the language in excel file '''
	col = search_lang(csv_lang)

	''' Creating folders for each lanuage with version '''
	try:
		if not os.path.exists(folder_name):
			os.mkdir(folder_name)
			os.chdir(folder_name)
			lang_folder = os.getcwd()
		else:
			os.chdir(folder_name)
			lang_folder = os.getcwd()

	except OSError:
		print ('Error: Creating directory.' + folder_name)
	
	''' Changing the directory '''
	os.chdir("..")
	
	''' open and read the content from csv file '''
	file = open(csv_file, 'r', encoding = 'utf-8')
	creader = csv.reader(file)
	x = 0
	pre_book = ""
	pre_chapter = ""
	content = ""

	for rows in creader:
		for row in rows:
			''' Below regex to split the BCV code to book, chapter and verse number's '''
			search_bcv = re.match(r"(\d+)(\d{3})(\d{3})",row)
			if search_bcv:
				book_number = search_bcv.group(1)
				chapter_number = search_bcv.group(2)
				verse_number = search_bcv.group(3)

				''' Pre-book is being checked for avoiding duplications '''
				if str(book_number) != str(pre_book) or x == 0:
					
					if content:

						''' Creating a usfm file with the proper content and file name '''
						final_content = re.sub(r"\*","",content)
						os.chdir(lang_folder)
						usfm_obj = open( bk_name[1] + ".usfm", 'w', encoding = "utf-8")
						usfm_obj.write(final_content)
						os.chdir("..")
						content = ""
						final_content = ""
					
					''' Add the required tag's and content '''
					bk_name = search_name(book_number,col)
					content += "\\id " + bk_name[1] + "\n" + "\\h " + bk_name[0] + "\n"
					content += "\\toc1 " + bk_name[0] + "\n" + "\\toc2 " + bk_name[0] + "\n" + "\\mt" + bk_name[0] + "\n"
					pre_book = book_number
					pre_chapter = ""
					x = 1
					y = 0

				''' Pre-chapter is being check for avoiding the duplications '''
				if str(chapter_number) != str(pre_chapter) or y == 0:
					content += "\\c " + chapter_number.lstrip("0") + "\n"
					print(chapter_number.lstrip("0"))
					pre_chapter = chapter_number
					y = 1

				''' The lstrip is used for avoiding the the extra zero's from the numbers '''
				content += "\\v " + verse_number.lstrip("0") + " "
			
			elif row:
				content += row + "\n"
				print(row)

	''' This block is for creating the usfm for last file in the array'''
	if content:
		final_content = re.sub(r"\*","",content)
		os.chdir(lang_folder)
		usfm_obj = open( bk_name[1] + ".usfm", 'w', encoding = "utf-8")
		usfm_obj.write(final_content)
		os.chdir("..")
		content = ""
		final_content = ""
'''
Below code is for testing by creating a single file.
'''
'''
	os.chdir(lang_folder)
	f_c = re.sub(r"\*","",content)
	txt_obj = open( csv_lang + ".txt", 'w', encoding = "utf-8")
	txt_obj.write(f_c)
	os.chdir("..")
'''