'''
Generate a csv file with data in an excel file.
'''

from openpyxl import load_workbook
import openpyxl
import csv
import os
import pandas as pd

''' Give the name of excel file '''
excel_file = "Bible_Summaries.xlsx"

if os.path.exists(excel_file):
	work_book = load_workbook(excel_file)
	sheets = work_book.get_sheet_names()
	work_sheet_b = work_book.get_sheet_by_name(sheets[0])
	work_sheet_c = work_book.get_sheet_by_name(sheets[1])
	work_book.active

	'''Creating a csv file'''
	with open ('Bible_Summaries.tsv', 'w', encoding = 'utf-8') as csv_file:
		cwriter = csv.writer(csv_file,dialect='excel')
		
		''' Adding column names in csv'''
		cwriter.writerow(["Book","Number_of_chapters","Key_Thought","Key_Verse","Christ_seen_as","Writer","Date",
			"Chapter_content","Character","Key_word","Striking_Facts","Conclusion"]);


		row = 2
		for b_row in range(2,work_sheet_b.max_row + 1):
			book = work_sheet_b['C' + str(b_row)].value
			
			if book:
				print("-----------",book,b_row)
				book_number = work_sheet_b['B' + str(b_row)].value
				b_total_chapter = work_sheet_b['E' + str(b_row)].value
				b_key_thought = work_sheet_b['D' + str(b_row)].value
				b_key_verse = work_sheet_b['F' + str(b_row)].value
				b_christ_as = work_sheet_b['G' + str(b_row)].value
				b_writer = work_sheet_b['H' + str(b_row)].value
				b_date = work_sheet_b['I' + str(b_row)].value
				b_conclusion = work_sheet_b['J' + str(b_row)].value
				bcv_code = str(book_number).zfill(2) + "0".zfill(6)

				''' Writing the book content into csv file column vise'''
				cwriter.writerow([bcv_code,b_total_chapter,b_key_thought,b_key_verse,b_christ_as,b_writer,b_date," "," "," "," ",b_conclusion]);
				#print(book,int(b_total_chapter),b_key_thought,b_key_verse,b_christ_as,b_writer,b_date,b_conclusion)

				for c_row in range(row,work_sheet_c.max_row + 1):
					c_book = work_sheet_c['B' + str(c_row)].value
					print("********",c_book,c_row)
					if c_book:
						#print(book.lower() + "====" + c_book.lower())
						if str(book.lower()) == str(c_book.lower()):
							c_chapter = work_sheet_c['C' + str(c_row)].value
							c_content = work_sheet_c['D' + str(c_row)].value
							c_character = work_sheet_c['E' + str(c_row)].value
							c_key_word = work_sheet_c['G' + str(c_row)].value
							c_strong_verse = work_sheet_c['H' + str(c_row)].value
							c_striking_fact = work_sheet_c['I' + str(c_row)].value
							c_conclusion = work_sheet_c['F' + str(c_row)].value
							bc_code = str(book_number).zfill(2) + str(c_chapter).zfill(3) + "0".zfill(3)
							print(book,c_content)

							'''Writing into csv file'''
							cwriter.writerow([bc_code,"","",c_strong_verse,"","","",c_content,c_character,c_key_word,c_striking_fact,c_conclusion]);
						else:
							break

				row = c_row