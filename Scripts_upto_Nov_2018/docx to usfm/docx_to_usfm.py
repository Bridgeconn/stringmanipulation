'''
Converting table docx files to usfm format by adding the tags.
'''

from docx import Document
import glob
import os
import re
from openpyxl import load_workbook
import openpyxl

def search_name(book):
	'''Directory change'''
	os.chdir("..")	
	excel_file = "BibleCodes.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for row in range(2,worksheet.max_row + 1):	
			book_name = worksheet['A' + str(row)].value
			
			'''Compares the book name with excel sheet and returns the short form'''
			if str(book) == str(book_name):
				short_name = worksheet['B' + str(row)].value
				os.chdir("docx")
				return short_name

'''Checking the docx directory and accessing it.'''
try:
	if not os.path.exists("docx"):
		print("No folder for docx, Create a folder with name \"docx\" and drop docx file in that")
	else:
		os.chdir("docx")
except OSError:
	print ('Error: Creating directory.' + "docx")

files = glob.glob("*.docx")

for docx_file in files:
	
	'''reading the contents from docx file'''
	document = Document(docx_file)
	final_content = ""
	content = ""
	x = 0

	file_name = re.sub(r".docx","",docx_file)

	'''Fetching the content'''
	for paragraph in document.paragraphs:
		para = paragraph.text
		chapter = re.sub(r"(अध्याय) (\d+)\.?","\\c \\2",para)
		
		'''To avoid blank lines'''
		if chapter:
			if chapter[0].isdigit():
				verse = "\\v " + chapter
			else:
				verse = chapter

			if chapter[0].isalpha():
				book_name = chapter.split("-")
				if x == 0:
					book = book_name[0]
					b = book.strip(" ")
					
					'''Calling function for book name'''
					rename = search_name(b)
					verse = "\\id " + rename
					x = 1

		content += verse + "\n"

	'''Directory change'''
	os.chdir("..")	

	'''Creating a folder or accessing a folder for converted usfm files'''
	try:
		if not os.path.exists("usfm"):
			os.mkdir("usfm")
			os.chdir("usfm")
		else:
			os.chdir("usfm")
	except OSError:
		print ('Error: Creating directory.' + "usfm")
	
	'''Writing into usfm file'''
	txt_file = open(file_name + ".usfm", 'w')
	txt_file.write(content)
	print(content)

	os.chdir("..")