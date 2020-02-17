''' 
Fetch the file content of particular column from one excel file to store it into the
different columns of another file with the same name, both the files are in different folders.

Write the translation words from one excel file to main excel file
with aligning it. The script is for multiple files.
'''

import glob
import openpyxl
import os
import re

def chapters(main_sheet, trans_sheet, book_main, file_path, previous_chapter):
	notes = ""
	for row in range(2, trans_sheet.max_row + 1):
		trans_chapter = trans_sheet['B' + str(row)].value
		
		''' Conditions are for checking the chapters and verse numbers from both the files '''
		if trans_chapter != None:
			trans_verse = trans_sheet['C' + str(row)].value


			if trans_verse != None:
				trans_versus = trans_sheet['D' + str(row)].value

				if trans_versus != None:
					trans_notes = trans_sheet['E' + str(row)].value

					if trans_notes != None:
						notes += "* " + str(trans_versus) + " - " + str(trans_notes) + "\n"

						#print(trans_chapter, trans_verse,notes)
						books(main_sheet, trans_chapter, trans_verse, trans_versus, notes, book_main, file_path, previous_chapter)
						notes = ""
						trans_verse = ""
						trans_chapter = ""

					else:
						notes += "* " + str(trans_versus) + "\n"
				else:
					trans_notes = trans_sheet['E' + str(row)].value
					
					if trans_notes != None:
						notes += str(trans_notes) + "\n"
					else:
						notes += ""

					#print(trans_chapter, trans_verse,notes)
					books(main_sheet, trans_chapter, trans_verse, trans_versus, notes, book_main, file_path, previous_chapter)
		else:
			trans_verse = trans_sheet['C' + str(row)].value

			if trans_verse != None:
				trans_versus = trans_sheet['D' + str(row)].value

				if trans_versus != None:
					trans_notes = trans_sheet['E' + str(row)].value

					if trans_notes != None:
						notes += "* " + str(trans_versus) + " - " + str(trans_notes) + "\n"

					else:
						notes += "* " + str(trans_versus) + "\n"
				else:
					trans_notes = trans_sheet['E' + str(row)].value

					if trans_notes != None:
						notes += str(trans_notes) + "\n"

					else:
						notes += ""

			else:
				trans_versus = trans_sheet['D' + str(row)].value

				if trans_versus != None:
					trans_notes = trans_sheet['E' + str(row)].value

					if trans_notes != None:
						notes += "* " + str(trans_versus) + " - " + str(trans_notes) + "\n"

					else:
						notes += "* " + str(trans_versus) + "\n"
				else:
					trans_notes = trans_sheet['E' + str(row)].value

					if trans_notes != None:
						notes += str(trans_notes) + "\n"

					else:
						notes += ""


def books(main_sheet, trans_chapter, trans_verse, trans_versus, remove_link, book_main, file_name, previous_chapter):
	
	note = re.sub(r"(\[.*?\]]+)?","",remove_link)
	no_link = re.sub(r"\[.*?\)","",note)
	notes = re.sub(r"(<b>.*<b>-?)?","",no_link)

	for row in range(2,main_sheet.max_row + 2):
		main_chapter = main_sheet['B' + str(row)].value

		if str(main_chapter) == str(trans_chapter):
			main_verse = main_sheet['C' + str(row)].value
			
			''' To Change single digit by adding '0' to it, if needed '''
			verse = main_verse.split("-")
			ver_num = ""
			a = ""
			for v_num in verse:
				a = ver_num
				ver_num = ""
				if len(v_num) < 2:
					ver_num = "0" + v_num
				else:
					ver_num = v_num

			if a:
				if ver_num:
					if a != ver_num:
						main_verse = a + "-" + ver_num
					else:
						main_verse = ver_num
				else:
					main_verse = a

			if len(main_verse) < 2:
				main_verse = "0" + main_verse
			''' To avoid the similar verse numbers from same file like 14-14 or 14-00 '''
			search_verse = re.match(r"(.\d)-(.\d)",trans_verse)
			
			if search_verse:
				if search_verse.group(1) == search_verse.group(2) or str(search_verse.group(2)) == "00":
					trans_verse = search_verse.group(1)
			
			''' Writing into target file (main file) '''
			if str(main_verse) == str(trans_verse):
				print("Hi=")
				print(main_chapter,trans_chapter, trans_verse,main_verse,notes)

				update_into(row,notes,main_sheet,book_main,file_name)
				main_chapter = ""
				main_verse = ""

		else:
			if trans_chapter:
				
				if str(main_chapter) == str(trans_chapter-1):
					main_verse = main_sheet['C' + str(row)].value
					first_verse = re.match(r"(.?\d)-?(.\d)?", trans_verse)
					
					if first_verse:
						x = first_verse.group(1)

					search_main = re.match(r"(.?\d)-?(.\d)?",main_verse)
					if search_main:
						y = search_main.group(1)

					if str(x) == str(y):
						check_chap = main_sheet['B' + str(row)].value
						if str(check_chap) == str(main_chapter):
							col_check = main_sheet['E' + str(row)].value
							if not col_check:
								print("Ho=")
								print(check_chap,main_chapter,trans_chapter, trans_verse,main_verse,notes)

								update_into(row,notes,main_sheet,book_main,file_name)
								main_chapter = ""
								trans_chapter = ""
								trans_verse = ""
								notes = ""
								x = ""
								y = ""

def update_into(row,notes,main_sheet,book_main,file_name):
	main_sheet['E' + str(row)] = notes
	book_main.save(file_name)



''' Fetch files from main folder '''
main_files = sorted(glob.glob("main/*.xlsx"))

for main_file in main_files:
	''' Fetch files from translation folder '''
	trans_files = sorted(glob.glob("translation/*.xlsx"))
	cwd = os.getcwd()
	file_path = cwd + "/" + main_file

	for trans_file in trans_files:

		''' Read main file -> Sheet '''
		book_main = openpyxl.load_workbook(main_file)
		main_sheets = book_main.get_sheet_names()
		main_sheet = book_main.get_sheet_by_name(main_sheets[0])
		book_main.active

		''' Read translation file -> Sheet '''
		book_trans = openpyxl.load_workbook(trans_file)
		trans_sheets = book_trans.get_sheet_names()
		trans_sheet = book_trans.get_sheet_by_name(trans_sheets[0])
		book_trans.active

		''' Get the book name from both the files '''
		m_book_name = main_sheet['A' + str(2)].value
		t_book_name = trans_sheet['A' + str(2)].value

		try:
			''' Checking the book name is same or not '''
			book_search = re.findall(m_book_name.lower(), t_book_name.lower())
			previous_chapter = ""
			if book_search:
				chapters(main_sheet, trans_sheet, book_main, file_path, previous_chapter)
		except:
			print(m_book_name)