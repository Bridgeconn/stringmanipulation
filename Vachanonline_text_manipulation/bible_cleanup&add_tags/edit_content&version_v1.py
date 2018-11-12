''' 
Script is for removing the unwanted tags, content and adding the bible book names in 
same language for multiple usfm files in different language folder.
'''

import os
import glob
import re
from openpyxl import load_workbook
import openpyxl

''' Function is for searching the language of the file '''
def search_lang(lang,loc):
	os.chdir("..")

	''' This excel file consist of the bible book name's in various language '''
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for column in "BCDEFGHIJKLMN":
			cell_name = "{}{}".format(column, 1)
			language = worksheet[cell_name].value

			''' It compare(s) the language with the .xlsx file '''
			if language.lower() == lang.lower():
				print(language,lang)
				os.chdir(loc)
				return column


''' Function is for updating the book name in same language '''
def search_name(path,book,col):
	os.chdir("..")
	os.chdir("..")
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for row in range(2,worksheet.max_row + 1):	
			book_name = worksheet["A" + str(row)].value
			
			''' Compares the book name with excel sheet 
				and returns the book name in same languag '''
			if str(book) == str(book_name):
				book_name = worksheet[col + str(row)].value
				os.chdir(path)
				return book_name


root = sorted(glob.glob("*/"))

for folder in root:
	lang = ""
	os.chdir(folder)
	loc = os.getcwd()

	''' Extract the language name from the folder name '''
	lang = re.sub(r"()(-.*)","",folder)
	
	''' Search the language in excel file '''
	col = search_lang(lang,loc)

	os.chdir("..")

	''' Creating folders for each lanuage '''
	folder_name = lang.lower() + "_irv"
	try:
		if not os.path.exists(folder_name):
			os.mkdir(folder_name)
			os.chdir(folder_name)
			new_folder = os.getcwd()
		else:
			os.chdir(folder_name)
			new_folder = os.getcwd()

	except OSError:
		print ('Error: Creating directory.' + folder_name)
	
	''' Changing the directory '''
	os.chdir(loc)
	stage_folder = glob.glob("*")
	
	for stage in stage_folder:
		
		os.chdir(stage)
		file_ext = sorted(glob.glob("*"))
		
		for file in file_ext:
			obj_file = open(file,'r',encoding = "utf-8")
			read_file = obj_file.read()
			
			''' File is being deleted after fetching the contents '''
			usfm_file = read_file.split("\n")
			search_id =  ""
			path = os.getcwd()
			
			''' Searching the id tag for book name '''
			for line in usfm_file:
				search_id = re.match(r"(\\id )(...)",line)
				if search_id:
					book = search_id.group(2)
					break

			''' Function call for getting book name's in same language '''
			new_name = search_name(path,book,col)
			print(new_name)

			''' Unwanted lines are being removed '''
			remove_mt = re.sub(r"\\mt.*","",read_file)
			remove_toc = re.sub(r"\\toc.*","",remove_mt)
			file_content = re.sub(r"\\h.*","",remove_toc)
			content = re.sub(r"\n+","\n",file_content)

			notes = ""
			doc = content.split("\n")
			x = 0
			
			''' Adding the data(s) into usfm file and cleaning up of unwanted data'''
			for line_each in doc:
				if line_each:
					main_content = ""
					line1 = ""
					a = line_each.split("\\f\*")
			
					for b in a:
						b = b + "\n"
						line1 += b
						b = ""
					
					'''This is used for avoiding the non space issue in verse'''
					add_space = re.match(r"(\\v \d+)(.*)",line1)
					if add_space:
						line1 = add_space.group(1) + " " + add_space.group(2)

					search_footnote = re.sub(r"(\\f[\s\+\\\w+\d+\:\-\.\W+]*\\f\*)","",line1)
					line2 = re.sub(r"\\\\f","",search_footnote)
					line3 = re.sub(r"\(.*\)","",line2)
					line4 = re.sub(r"(\s)+"," ",line3)
					rows = re.sub(r"\n","",line4)
					search_tag = ""
					array_tag = ["\\imt","\\im","\\ip","\\ide","\\io","\\iot","\\is","\\ior","\\ior\*"]
					for i in array_tag:
						search_tag = rows.find(i)
						if search_tag != -1:
							print(search_tag)
							break
					
					#print(rows)
					search_ref = re.sub(r"\\r(.*)","",rows)
					search_ids = re.match(r"(\\id )(...)",search_ref)
					if search_ids:
						if x == 0:
							print("*****"+new_name)
							notes += "\\id " + search_ids.group(2) + "\n\\h " + new_name + "\n\\toc1 " + new_name + "\n\\toc2 " + new_name + "\n\\mt " + new_name + "\n"
							x = 1

					elif search_tag:
						notes += search_ref + "\n"
						#print(search_ref)

			''' Creating a new file with same name and adding final data
				in the language folder '''
			os.chdir(new_folder)
			extn_change = re.sub(r"\.(.*)",".usfm",file)
			new_file = open(extn_change,'w',encoding = "utf-8")
			notes1 = re.sub("[\<\>]","",notes)
			#notes1 = re.sub("(\\\*)","",notes2)
			new_file.write(notes1)
			print(notes1)
			os.chdir(path)

		os.chdir("..")

	os.chdir("..")