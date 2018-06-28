'''
	The script is written for generating the excel files 
	for Bible books(66) with a column with chapter number.
'''
import openpyxl
from openpyxl import load_workbook

'''Array containing chapter number with book standard name'''
ch_in_book = {"GEN": ["1", 50], "EXO": ["2", 40], "LEV": ["3", 27], 
  "NUM": ["4", 36], "DEU": ["5", 34], "JOS": ["6", 24], "JDG": ["7", 21], 
  "RUT": ["8", 4], "1SA": ["9", 31], "2SA": ["10", 24], "1KI": ["11", 22], 
  "2KI": ["12", 25], "1CH": ["13", 29], "2CH": ["14", 36], "EZR": ["15", 10], 
  "NEH": ["16", 13], "EST": ["17", 10], "JOB": ["18", 42], "PSA": ["19", 150], 
  "PRO": ["20", 31], "ECC": ["21", 12], "SNG": ["22", 8], "ISA": ["23", 66], 
  "JER": ["24", 52], "LAM": ["25", 5], "EZK": ["26", 48], "DAN": ["27", 12], 
  "HOS": ["28", 14], "JOL": ["29", 3], "AMO": ["30", 9], "OBA": ["31", 1], 
  "JON": ["32", 4], "MIC": ["33", 7], "NAM": ["34", 3], "HAB": ["35", 3], 
  "ZEP": ["36", 3], "HAG": ["37", 2], "ZEC": ["38", 14], "MAL": ["39", 4], 
  "MAT": ["40", 28], "MRK": ["41", 16], "LUK": ["42", 24], "JHN": ["43", 21], 
  "ACT": ["44", 28], "ROM": ["45", 16], "1CO": ["46", 16], "2CO": ["47", 13], 
  "GAL": ["48", 6], "EPH": ["49", 6], "PHP": ["50", 4], "COL": ["51", 4], 
  "1TH": ["52", 5], "2TH": ["53", 3], "1TI": ["54", 6], "2TI": ["55", 4], 
  "TIT": ["56", 3], "PHM": ["57", 1], "HEB": ["58", 13], "JAS": ["59", 5], 
  "1PE": ["60", 5], "2PE": ["61", 3], "1JN": ["62", 5], "2JN": ["63", 1], 
  "3JN": ["64", 1], "JUD": ["65", 1], "REV": ["66", 22]}

'''Here we call each data from the array'''
for bookname in ch_in_book:
	print(bookname)
	book = bookname + ".xlsx"
	
	'''Generating excel file with book name(array value)'''
	work_book = openpyxl.Workbook()
	work_book.save(book)
	work_book = load_workbook(book)
	worksheet = work_book.get_sheet_by_name("Sheet")
	work_book.active

	'''Adding the column heading'''
	worksheet['A1'] = "Chapter"
	worksheet['B1'] = "Page No."
	work_book.save(book)

	'''Fetching the chapter number from array'''
	row = ch_in_book[bookname]
	chapters = row[1]
	print(chapters)
	chapter = 1
	rows = 2

	while chapters >= chapter:
		print(chapter)
		
		'''Writing the values row by row'''
		worksheet['A' + str(rows)] = chapter
		chapter += 1
		rows += 1

	work_book.save(book)