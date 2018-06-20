''' 
Script is for removing the unwanted tags and adding the bible book names in 
same language for multiple usfm files in different language folder.
'''

import os
import glob
import re
from openpyxl import load_workbook
import openpyxl

''' Function is for searching the language of the file '''
def search_lang(lang,loc):
	os.chdir("..")

	''' This excel file consist of the bible book name's in various language '''
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for column in "BCDEFGHIJKLMN":
			cell_name = "{}{}".format(column, 1)
			language = worksheet[cell_name].value
			#print(language)

			''' It compare(s) the language with the .xlsx file '''
			if language.lower() == lang.lower():
				print(language,lang)
				os.chdir(loc)
				return column


''' Function is for updating the book name in same language '''
def search_name(path,book,col):
	os.chdir("..")
	os.chdir("..")
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for row in range(2,worksheet.max_row + 1):	
			book_name = worksheet["A" + str(row)].value
			
			''' Compares the book name with excel sheet 
				and returns the book name in same language '''
			if str(book) == str(book_name):
				book_name = worksheet[col + str(row)].value
				#os.chdir("docx")
				os.chdir(path)
				return book_name


root = glob.glob("*/")

for folder in root:
	lang = ""
	os.chdir(folder)
	loc = os.getcwd()

	''' Extract the language name from the folder name '''
	lang = re.sub(r"()(-.*)","",folder)
	
	''' Search the language in excel file '''
	col = search_lang(lang,loc)

	os.chdir("..")

	''' Creating folders for each lanuage '''
	folder_name = lang.lower() + "_irv"
	try:
		if not os.path.exists(folder_name):
			os.mkdir(folder_name)
			os.chdir(folder_name)
			new_folder = os.getcwd()
		else:
			os.chdir(folder_name)
			new_folder = os.getcwd()

	except OSError:
		print ('Error: Creating directory.' + folder_name)
	
	''' Changing the directory '''
	os.chdir(loc)
	stage_folder = glob.glob("*")
	
	for stage in stage_folder:
		if stage.lower() == "Stage 3".lower():
			os.chdir(stage)
			file_ext = glob.glob("*")
			
			for file in file_ext:
				obj_file = open(file,'r',encoding = "utf-8")
				read_file = obj_file.read()
				
				''' File is being deleted after fetching the contents '''
				#os.remove(file)
				usfm_file = read_file.split("\n")
				search_id =  ""
				path = os.getcwd()
				
				''' Searching the id tag for book name '''
				for line in usfm_file:
					search_id = re.match(r"(\\id )(...)",line)
					if search_id:
						book = search_id.group(2)
						break

				''' Function call for getting book name's in same language '''
				new_name = search_name(path,book,col)

				#print(read_file)
				''' Unwanted lines are being removed '''
				remove_mt = re.sub(r"\\mt.*","",read_file)
				remove_toc = re.sub(r"\\toc.*","",remove_mt)
				file_content = re.sub(r"\\h.*","",remove_toc)
				content = re.sub(r"\n+","\n",file_content)

				#print(content)
				notes = ""
				doc = content.split("\n")
				x = 0
				
				''' Adding the data(s) into usfm file '''
				for rows in doc:
					search_tag = ""
					array_tag = ["\\imt","\\im","\\ip","\\io","\\iot","\\ior","\\ior\*"]
					#search_tag = re.match(r"(\\imt)?(\\im)?(\\ip)?(\\io)?(\\iot)?(\\ior)?",rows)
					for i in array_tag:
						search_tag = rows.find(i)
						if search_tag != -1:
							print(search_tag)
							break
					search_ids = re.match(r"(\\id )(...)",rows)
					if search_ids:
						if x == 0:
							#notes += rows
							#print(rows,new_name)
							notes += rows + "\n\\h " + new_name + "\n\\toc1 " + new_name + "\n\\toc2 " + new_name + "\n\\mt " + new_name + "\n"
							x = 1

					elif search_tag:
						#print(rows)
						notes += rows + "\n"

				''' Creating a new file with same name and adding final data
					in the language folder '''
				os.chdir(new_folder)
				extn_change = re.sub(r"\.(.*)",".usfm",file)
				new_file = open(extn_change,'w',encoding = "utf-8")
				new_file.write(notes)
				print(notes)
				os.chdir(path)

			os.chdir("..")

	os.chdir("..")