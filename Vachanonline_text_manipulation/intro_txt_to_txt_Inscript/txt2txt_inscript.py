'''
	The script is for generating a txt file with the content where a line will have the content of book,
	The input and output files are txt files.
'''
# -*- encoding: utf-8 -*-

from docx import Document
import glob
import os
import re
import sys
from openpyxl import load_workbook
import openpyxl


'''Function is for searching the language of the file'''
def search_lang_col(lang):
	os.chdir("..")
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for column in "DEFGHIJKLMNOPQR":
			cell_name = "{}{}".format(column, 1)
			language = worksheet[cell_name].value.split("/")

			''' It compare(s) the language with the .xlsx file '''
			if language[1].lower() == lang.lower():
				print(language[1],lang)
				os.chdir("Source")
				return column

''' Function is for updating the book name in same language '''
def search_name(books,col):
	book = books.rstrip(" ")
	os.chdir("..")
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for row in range(2,worksheet.max_row + 1):	
			books_name = worksheet[col + str(row)].value
			
			''' Compares the book name with excel sheet 
			and returns the book name in same languag '''
			search_b_n = re.match(books_name,book)
			if search_b_n:
				book_name = worksheet[col + str(row)].value
				book_id = worksheet["B" + str(row)].value
				os.chdir("Source")
				return book_name,book_id


'''Creating a target directory for storing the generating usfm files'''
try:
	if not os.path.exists("Target"):
		os.mkdir("Target")
		os.chdir("Target")
		path = os.getcwd()
		os.chdir("..")

	else:
		os.chdir("Target")
		path = os.getcwd()
		os.chdir("..")

except OSError:
	print ('Error: No directory.' + "Target")
	sys.exit()


'''Checking the docx directory and accessing it.'''
try:
	if not os.path.exists("Source"):
		print("No folder for Source, Create a folder with name \"Source\" and drop Source file in that")
		sys.exit()
	else:
		os.chdir("Source")
		source_path = os.getcwd()
except OSError:
	print ('Error: No directory.' + "Source")
	sys.exit()



files = glob.glob("*.txt")

content = ""
for txt_file in files:
	
	'''Reading the contents from docx file'''
	file_name = re.sub(r".txt","",txt_file)
	search_lang = re.match(r"(.{3})_?(.*)(\.txt)",txt_file)
	if search_lang:
		file_name = search_lang.group(1)
		print(file_name)

	lang_col = search_lang_col(file_name)
	print(lang_col)

	'''Fetching the content'''
	txt_obj = open(txt_file, 'r', encoding = "utf-8")
	txt_content1 = txt_obj.read()
	txt_content = re.sub(r"	","",txt_content1)
	txt_lines = txt_content.split("\n")
	x = 1
	i = 0
	
	'''Reading each line from the file'''
	book_code = ""
	for txt_line in txt_lines:
		if txt_line:

			'''Searching book ID and other tags with content'''
			search_id = re.match(r"\\id (.*)",txt_line)
			remaining_tags = re.match(r"(\d+\.?\s?)(.*)(\(([\:\-\d+\s\–]*)\))",txt_line)
			if search_id:
				txt_id = search_id.group(1)
				book = search_name(txt_id.lstrip(" "),lang_col)
				book_code = book[1]
				content += "\n" + "\\id " + book[1] + "\t"
				i = 1

			elif len(txt_line) < 10:
				search_is_tag = re.match(r"लेखक",txt_line)
				if search_is_tag:
					content += "<br><b>" + txt_line + "</b> "
					i = 1

			''' Adding the various tags for the content '''
			tags = "प्रापक","लेखन तिथि एवं स्थान","उद्देश्य","रूपरेखा"

			for tag in tags:
				search_is_tag = re.match(tag,txt_line)
				if search_is_tag:
					content += "<br><b>" + txt_line + "</b> "
					i = 1
					break

			if remaining_tags:
				content += remaining_tags.group(2) + remaining_tags.group(3)

			elif (i == 0):
				content += txt_line

			i = 0

	os.chdir(source_path)

print(content)
'''Generating intro.txt file in the Target folder with the edited content for Inscript'''
if content:
	final_content = re.sub(r"\\id ","",content)
	file = open( path + "/" + "intro" + ".txt", 'w', encoding = "utf-8")
	file.write(final_content)