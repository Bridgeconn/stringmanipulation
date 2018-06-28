'''
 Converting single .txt file to  multiple .usfm file by adding tags and more useful contents.
 the script is for adding the introduction to each books.
'''
# -*- encoding: utf-8 -*-

from docx import Document
import glob
import os
import re
import sys
from openpyxl import load_workbook
import openpyxl


'''Function is for searching the language of the file'''
def search_lang_col(lang):
	os.chdir("..")
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for column in "CDEFGHIJKLMNOPQ":
			cell_name = "{}{}".format(column, 1)
			language = worksheet[cell_name].value.split("/")

			''' It compare(s) the language with the .xlsx file '''
			if language[1].lower() == lang.lower():
				print(language[1],lang)
				os.chdir("Source")
				return column

''' Function is for updating the book name in same language '''
def search_name(books,col):
	book = books.rstrip(" ")
	os.chdir("..")
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for row in range(2,worksheet.max_row + 1):	
			books_name = worksheet[col + str(row)].value

			''' Compares the book name with excel sheet 
			and returns the book name in same languag '''
			search_b_n = re.match(books_name,book)
			if search_b_n:
				book_name = worksheet[col + str(row)].value
				book_id = worksheet["B" + str(row)].value
				os.chdir("Source")
				return book_name,book_id


'''Creating a target directory for storing the generating usfm files'''
try:
	if not os.path.exists("Target"):
		os.mkdir("Target")
		os.chdir("Target")
		path = os.getcwd()
		os.chdir("..")

	else:
		os.chdir("Target")
		path = os.getcwd()
		os.chdir("..")

except OSError:
	print ('Error: No directory.' + "Target")
	sys.exit()


'''Checking the docx directory and accessing it.'''
try:
	if not os.path.exists("Source"):
		print("No folder for Source, Create a folder with name \"Source\" and drop Source file in that")
		sys.exit()
	else:
		os.chdir("Source")
except OSError:
	print ('Error: No directory.' + "Source")
	sys.exit()



files = glob.glob("*.txt")

for txt_file in files:
	
	'''Reading the contents from docx file'''
	file_name = re.sub(r".txt","",txt_file)
	search_lang = re.match(r"(.{3})_?(.*)(\.txt)",txt_file)
	if search_lang:
		file_name = search_lang.group(1)
		print(file_name)

	lang_col = search_lang_col(file_name)
	print(lang_col)

	'''Fetching the content'''
	txt_obj = open(txt_file, 'r', encoding = "utf-8")
	txt_content = txt_obj.read()
	txt_lines = txt_content.split("\n")
	x = 1
	i = 0
	content = ""
	book_code = ""

	'''Reading each line'''
	for txt_line in txt_lines:
		if txt_line:
			search_id = re.match(r"\\id (.*)",txt_line)
			remaining_tags = re.match(r"(\d+\.?\s?)(.*)(\(([\:\-\d+\s\–]*)\))",txt_line)
			if search_id:
				
				if content:
					''' If content has some values then it will create the usfm file with compplete tags and content '''
					print(book_code,content,x)
					usfm_file = open(path + "/" +book_code + ".usfm", 'w', encoding = "utf-8")
					usfm_file.write(content)
					content = ""
					book_code = ""
					x += 1
				
				'''Using the function (search_name) we get the book 3 letter standard name and
				 book name in same language as in excel'''
				txt_id = search_id.group(1)
				book = search_name(txt_id.lstrip(" "),lang_col)
				book_code = book[1]
				content += "\\id " + book[1] + "\n" + "\\mt " + book[0] + "\n"
				x += 1
				i = 1

			elif len(txt_line) < 10:
				search_is_tag = re.match(r"लेखक",txt_line)
				if search_is_tag:
					content += "\\imt " + txt_line + "\n"
					i = 1

			''' Adding the various tags for the content '''
			tags = "प्रापक","लेखन तिथि एवं स्थान","उद्देश्य"

			for tag in tags:
				search_is_tag = re.match(tag,txt_line)
				if search_is_tag:
					content += "\\is " + txt_line + "\n"
					i = 1
					break

			apply_iot_tag = re.match(r"रूपरेखा",txt_line)
			if apply_iot_tag:
				content += "\\iot " + txt_line + "\n"

			elif remaining_tags:
				content += "\\io " + remaining_tags.group(2) + "\\ior " + remaining_tags.group(3) + "\\ior*" + "\n"

			elif (i == 0):
				content += "\\ip " + txt_line + "\n"

			i = 0

	'''Creating the usfm file for each book with edited content'''
	if content:
		print(book_code,content,x)
		usfm_file = open( path + "/" + book_code + ".usfm", 'w', encoding = "utf-8")
		usfm_file.write(content)
		content = ""
		book_code = ""
		x+=1
