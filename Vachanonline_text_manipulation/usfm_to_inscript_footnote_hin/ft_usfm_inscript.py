'''
	The script is for collecting the data from multiple usfm files and generate a text file 
	by adding tags and editing the content. The book 3 letter standard name also takes from
	the excel file for Inscript App. 
'''

import glob
import os
import re
from openpyxl import load_workbook
import openpyxl

''' Function is for fetching the books 3 letter standard name for Inscript '''
def search_name(books,cwd):
	book = books.rstrip(" ")
	os.chdir("..")
	excel_file = "Book name 12 GL.xlsx"
	
	if os.path.exists(excel_file):
		work_book = load_workbook(excel_file)
		worksheet = work_book.get_sheet_by_name("Sheet1")
		work_book.active

		for row in range(2,worksheet.max_row + 1):	
			books_name = worksheet["C" + str(row)].value
			
			''' Compares the 3 letter paratext standard book name with excel sheet 
			and returns the 3 letter Inscript standard book name in same languag '''
			search_b_n = re.match(books_name,book)
			if search_b_n:
				book_id = worksheet["B" + str(row)].value
				os.chdir(cwd)
				return book_id

''' Access the folder with the name "usfm" '''
folder = "usfm"
os.chdir(folder)
files = glob.glob("*.usfm")
main_content = ""

for file in files:
	cwd = os.getcwd()
	print(file,cwd)

	''' Reads the usfm file '''
	usfm_file = open(file, 'r')
	content = usfm_file.read()
	lines = content.split("\n")

	for line in lines:
		if line:
			
			'''Regex fetching the required data and adding required tags, datas '''
			search_book = re.match(r"(\\id)\s?(...)",line)
			search_chapter = re.match(r"(\\c)\s?(\d+)",line)
			search_verse = re.match(r"(\\v)\s?(.*)",line)
			search_quotes = re.match(r"(\\f)\s?(.*)\: (.*)",line)

			if search_book:
				file_name1 = search_book.group(2)
				''' Using paratext book code the function (search_name) is being called '''
				file_name = search_name(file_name1,cwd)
				#print(file_name)
			
			elif search_chapter:
				chapter = search_chapter.group(2)
				main_content += "\n" + file_name + " " + chapter + ":" + "\t" + "<b>Overview</b><br>"
				#print("\n" + file_name + " " + chapter + ":" + "\t" + "<b>Overview</b><br>")

			elif search_verse:
				verse = search_verse.group(2)
				main_content += "<u>" + file_name + "_" + chapter+":"+ verse+ " </u>,"
				#print("<u>" + file_name + "_" + chapter+":"+ verse+ " </u>,")

			elif search_quotes:
				quote = search_quotes.group(2)
				main_content += " <b> " + quote + ": </b>" + search_quotes.group(3) + " <br>"
				#print(" <b> " + quote + ": </b>" + search_quotes.group(3) + " <br>")

os.chdir("..")
print(main_content)

''' Writing the whole content into 1 text file named as footnote.txt'''
txt_file = open("footnote.txt",'w',encoding = "utf-8")
txt_file.write(main_content)