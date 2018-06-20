'''
 Fetching the Default values from xml file and writing it to excel file
 '''
 

import glob
import re
from openpyxl import load_workbook
import os
import openpyxl

files = glob.glob("*.xml")
cwd = os.getcwd()
txt = ""

for xml_file in files:
	file = open(xml_file,"r")
	file_name = re.sub(r".xml","",xml_file)
	excel_file = cwd + "/" + file_name + ".xlsx"
	print(excel_file)
	row = 2
	
	for line in file:
		search_def = re.match(r"(<DefaultValue>)(.*)?(<.DefaultValue>)",line,re.M|re.I)
		
		if search_def:
			txt = search_def.group(2)

			if os.path.exists(excel_file):
				'''Loading the existing file'''
				work_book = load_workbook(excel_file)
				worksheet = work_book.get_sheet_by_name("Sheet")	
				
				'''Updating the sheet with value'''
				worksheet['A' + str(row)] = txt
				work_book.save(excel_file)
				print(txt)
				row+=1
					
			else:
				'''Loading a new excel file with the same name of xml'''
				work_book = openpyxl.Workbook()
				work_book.save(excel_file)
				work_book = load_workbook(excel_file)
				worksheet = work_book.get_sheet_by_name("Sheet")	
				
				'''Updating the sheet with column header'''
				worksheet['A1'] = 'Default Value'
				worksheet['B1'] = 'Translation'
				
				'''Updating the sheet with values'''
				worksheet['A' + str(row)] = txt
				work_book.save(excel_file)
				print(txt)
				row+=1