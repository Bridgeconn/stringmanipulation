''''
 Script is used to find the word count of each .md file and total word count from the folder too.
'''


import glob
import os
import re
from os.path import join
from openpyxl import load_workbook
import openpyxl

cwd = os.getcwd()
content = []
totalcount = 0

'''Name of the folder consist of .md files'''
path = 'bible'

for root, sdirs, files in os.walk(path):
	for filename in files:
		file_path = os.path.join(root, filename)
		if file_path.endswith('.md'):
			with open(file_path) as f :
				'''p contains contents of entire file'''
				p = f.read()
				
				'''logic to compute word counts follows here.'''
				process3 = re.sub("\d+", "",p)
				process4 = re.sub("-", "",process3)
				process5 = re.sub("\(?\)?\[?\]?\.?\_?\*?\,?\??\!?\<?\>?\|?", "",process4)
				process6 = re.sub("#{1,2}","",process5)
				process7 = re.sub(":", "", process6)
				process8 = re.sub("\"", "", process7)
				words = process8.split()
				#print(words)

				filecount = len(words)
		#print filename,filecount
		
		'''Open a .txt file'''
		txt_file = open('bible.txt','w+')
		file_name = re.sub(".md","",filename)
		
		''' Stores the count of words and file name'''
		content.append(file_name + " = " + str(filecount))
		totalcount += filecount
		filecount = 0

#x = sorted(content)
'''Write the count along with the file name in txt file with total count too.'''
#content1 = strx + "\n" + "The total word count is:" + str(totalcount)
#txt_file.write(x)

x  = sorted(content)
txt_file.write(x)

excel_file = cwd + "/bible.xlsx"

if os.path.exists(excel_file):
	work_book = load_workbook(excel_file)
	worksheet = work_book.get_sheet_by_name("Sheet")
	work_book.active

	'''Writing the values row by row'''
	rows = 2
	for row in x:
		sep = re.match(r"(.*) = (\d+)",row)
		if sep:
			name = sep.group(1) + "\n"
			count = sep.group(2) + "\n"
			print(name,count)
				
			worksheet['A' + str(rows)] = name
			worksheet['B' + str(rows)] = count
			rows += 1
		work_book.save(excel_file)
else:
	work_book = openpyxl.Workbook()
	work_book.save(excel_file)
	work_book = load_workbook(excel_file)
	worksheet = work_book.get_sheet_by_name("Sheet")
	
	'''Updating the sheet with column'''
	worksheet['A1'] = 'File name'
	worksheet['B1'] = 'Word count'
	work_book.save(excel_file)

	'''Writing the values row by row to excel'''
	rows = 2
	for row in x:
		sep = re.match(r"(.*) = (\d+)",row)
		if sep:
			name = sep.group(1)
			count = sep.group(2)
			print(name,count)
				
			worksheet['A' + str(rows)] = name
			worksheet['B' + str(rows)] = count
			rows += 1
		work_book.save(excel_file)