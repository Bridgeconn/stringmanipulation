'''
To create a .docx file with table by comparing .usfm file and excel file
'''

import glob
import os
from openpyxl import load_workbook
import openpyxl
import re
from docx import Document

'''To create a docx folder, if it didn't exists''' 
def docx_folder():
	os.chdir("..")
	try:
		if not os.path.exists("docx"):
			os.makedirs("docx")
			os.chdir("docx")
		else:
			os.chdir("docx")
	except OSError:
		print ('Error: Creating directory.' + "docx")

'''The name of the folder which contains the usfm files'''
usfm_folder = "usfm"
os.chdir(usfm_folder)
usfm_files = sorted(glob.glob("*.usfm"))

for usfm_file in usfm_files:
	cwd = os.getcwd()

	'''Reads the usfm file'''
	print(cwd)
	open_usfm = open(usfm_file, 'r')
	usfm_content = open_usfm.read()
	usfm_lines = usfm_content.split("\n")

	'''To extract the usfm file name fromm the file'''
	for usfm_line in usfm_lines:
		search_id = re.match(r"(\\id\s)(.*)",usfm_line)
		if search_id:
			usfm_name = search_id.group(2)
			print(usfm_line,usfm_name)

	os.chdir("..")

	'''Working on excel file'''
	excel_folder = "excel"
	os.chdir(excel_folder)
	excel_files = sorted(glob.glob("*.xlsx"))
	
	for excel_file in excel_files:
		cwd = os.getcwd()
		quote_verse = ""
		tag_verse = ""
		add_verse = ""
		print(excel_file)

		'''Reads the excel file'''
		workbook = load_workbook(excel_file)
		worksheets = workbook.get_sheet_names()
		worksheet = workbook.get_sheet_by_name(worksheets[0])
		workbook.active
		
		excel_name = worksheet['A' + str(2)].value
		print(usfm_name, excel_name)

		usfm_full = ""
		book_search = re.findall(excel_name, usfm_name)
		
		if book_search:
			print(worksheet.max_row)

			'''Create a docx file'''
			docx_folder()
			docx_file = excel_name + ".docx"
			document = Document()
			if not docx_file:
				document.save(docx_file)

			'''Create a table in docx file'''
			table = document.add_table(rows = worksheet.max_row + 1, cols = 5)
			t_row = table.rows[0]
			t_row.cells[0].text = "Book"
			t_row.cells[1].text = "Chapter"
			t_row.cells[2].text = "Verse"
			t_row.cells[3].text = "Source"
			t_row.cells[4].text = "Target"
			table.style = "TableGrid"
			document.save(docx_file)
			i = 1
			
			usfm_chapter = ""
			usfm_verse = ""
			versus = ""

			for usfm_line in usfm_lines:
				if usfm_line:		

					'''Regex to remove tags and unwanted lines'''
					search_chapter = re.match(r"(\\c)\s(\d+)",usfm_line)
					search_verse = re.match(r"(\\q)?(\d+)?\s?(\\v)\s?(\d+)-?(\d+)?\s?(.*)",usfm_line)
					search_quotes = re.match(r"(\\q(\d+)?)\s?(.*)",usfm_line)
					search_tags = re.match(r"((\\\w+)\s?(\d+)?)?(.*)",usfm_line)

					'''Gets the chapter number'''
					if search_chapter:
						usfm_chapter = search_chapter.group(2)
						next_verse = ""
						print(usfm_chapter)
					
					'''Extract verse number and versus'''
					elif search_verse:
						usfm_verse = search_verse.group(4)
						next_verse = search_verse.group(5)
						versus = search_verse.group(6)
						add_verse = search_verse.group(6)
					
					'''Extract the content in quote's tag'''
					elif search_quotes:
						quote = search_quotes.group(3)
						if quote:
							row = table.rows[i-1]
							quote_verse += str(add_verse) + " " + str(quote)
							row.cells[3].text = str(quote_verse)
							document.save(docx_file)
							print(quote)
							add_verse = ""
							quote = ""
							row = ""

					'''Extract the content if any tags remaining'''
					elif search_tags:
						if add_verse or tag_verse:
							no_tags = search_tags.group(4)
							if no_tags:
								row_pre = table.rows[i-1]
								tag_verse += str(add_verse) + " " + str(no_tags)
								row_pre.cells[3].text = str(tag_verse)
								document.save(docx_file)
								print(no_tags)
								add_verse = ""
								no_tags = ""
								row = ""

					x = 0
					y = 0

					'''Reads the content from excel file'''
					for row in range(2,worksheet.max_row + 1):	
						excel_chapter = worksheet['C' + str(1)].value
						
						if x == 0:
							if excel_chapter.lower() == "chapter":
								chapter_num = worksheet['C' + str(row)].value
								x = 1
								cols = 'C'
							else:
								chapter_num = worksheet['B' + str(row)].value
								x = 1
								cols = 'B'
						else:
							chapter_num = worksheet[cols + str(row)].value

						excel_verse = worksheet['D' + str(1)].value

						if y == 0:
							if excel_verse.lower() == "verse":
								verse_num = worksheet['D' + str(row)].value
								y = 1
								col = 'D'
							else:
								verse_num = worksheet['C' + str(row)].value
								y = 1
								col = 'C'
						else:
							verse_num = worksheet[col + str(row)].value

						book = worksheet['A' + str(row)].value
						
						'''Write the content into docx file, (Fills the table)'''
						if str(chapter_num) == str(usfm_chapter):
							if  str(verse_num) == str(usfm_verse):
								quote_verse = ""
								tag_verse = ""
								row = table.rows[i]
								row.cells[0].text = str(book)
								row.cells[1].text = str(chapter_num)
								row.cells[2].text = str(usfm_verse)
								row.cells[3].text = str(versus)
								i += 1
								document.save(docx_file)
								
								print("*****",i,book,chapter_num,usfm_verse,versus)
								usfm_verse = ""
								versus = ""
								book = ""

							elif str(verse_num) == str(next_verse):
								row = table.rows[i]
								row.cells[0].text = str(book)
								row.cells[1].text = str(chapter_num)
								row.cells[2].text = str(next_verse)
								row.cells[3].text = ""
								i += 1
								document.save(docx_file)
								
								print("*****",i,book,chapter_num,next_verse,versus)
								next_verse = ""
								versus = ""
								book = ""
		os.chdir("..")
		os.chdir("excel")
	os.chdir("..")
	os.chdir("usfm")