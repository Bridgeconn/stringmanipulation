# Write the translation words from one excel file to main excel file
# with aligning it. The script is for multiple files


import glob
import openpyxl
import os
import re

def chapters(main_sheet, trans_sheet, book_main, file_path):
	notes = ""
	
	for row in range(2, main_sheet.max_row):
		main_chapter = main_sheet['B' + str(row)].value
		
		for rows in range(2, trans_sheet.max_row + 1):
			trans_chapter = trans_sheet['B' + str(rows)].value
			
			if len(main_chapter) < 2:
				main_chapter1 = "0" + main_chapter
			else:
				main_chapter1 = main_chapter

			if len(trans_chapter) < 2:
				trans_chapter1 = "0" + trans_chapter
			else:
				trans_chapter1 = trans_chapter

			if (main_chapter1 == trans_chapter1):

				main_verse = main_sheet['C' + str(row)].value
				trans_verse = trans_sheet['C' + str(rows)].value

				# To Change single digit by adding '0' to it, if needed 
				verse = main_verse.split("-")
				ver_num = ""
				a = ""
				for v_num in verse:
					a = ver_num
					ver_num = ""
					if len(v_num) < 2:
						ver_num = "0" + v_num
					else:
						ver_num = v_num

				if a and ver_num:
					if a != ver_num:
						main_verse = a + "-" + ver_num
					else:
						main_verse = ver_num

				# To Change single digit by adding '0' to it, if needed 
				verse = trans_verse.split("-")
				t_ver_num = ""
				b = ""
				for t_v_num in verse:
					b = t_ver_num
					t_ver_num = ""
					if len(t_v_num) < 2:
						t_ver_num = "0" + t_v_num
					else:
						t_ver_num = t_v_num

				if b and t_ver_num:
					if b != t_ver_num:
						trans_verse = b + "-" + t_ver_num
					else:
						trans_verse = t_ver_num
			

				if a:
					if b:
						#print("-----a-----",a,"=",b,"------b-----")
						if int(a) == int(b):
							print("a-###",a,"=",b,"###-b",rows)
							write_note(rows, row, main_sheet, trans_sheet, book_main, file_path)
					else:
						#print("-----a-----",a,"=",trans_verse,"-----trans_verse------")
						if a == trans_verse:
							print("a-###",a,"=",trans_verse,"###-trans_verse",rows)
							write_note(rows, row, main_sheet, trans_sheet, book_main, file_path)
				else:
					if b:
						#print("-----main_verse-----",main_verse,"=",b,"------b-----")
						if main_verse == b:
							print("main_verse-###",main_verse,"=",b,"###-b",rows)
							write_note(rows, row, main_sheet, trans_sheet, book_main, file_path)
					else:
						#print("-----main_verse-----",main_verse,"=",trans_verse,"------trans_verse-----")
						if main_verse == trans_verse:
							print("main_verse-###",main_verse,"=",trans_verse,"###-trans_verse",rows)
							write_note(rows, row, main_sheet, trans_sheet, book_main, file_path)

def write_note(rows, row, main_sheet, trans_sheet, book_main, file_path):

	trans_note = trans_sheet['D' + str(rows)].value
	main_sheet['E' + str(row)] = trans_note
	book_main.save(file_path)


# Fetch files from main folder
main_files = sorted(glob.glob("main/*.xlsx"))

for main_file in main_files:

	m_f_name = re.sub(r"main/","",main_file)
	
	# Fetch files from translation folder
	trans_files = sorted(glob.glob("add/*.xlsx"))
	cwd = os.getcwd()
	file_path = cwd + "/" + main_file

	for trans_file in trans_files:

		t_f_file = re.sub(r"add/","",trans_file)
	
		try:
			# Checking the book name is same or not
			book_search = re.findall(m_f_name.lower(), t_f_file.lower())
			
			if book_search:
				# Read main file -> Sheet
				book_main = openpyxl.load_workbook(main_file)
				main_sheets = book_main.get_sheet_names()
				main_sheet = book_main.get_sheet_by_name(main_sheets[0])
				book_main.active

				# Read translation file -> Sheet
				book_trans = openpyxl.load_workbook(trans_file)
				trans_sheets = book_trans.get_sheet_names()
				trans_sheet = book_trans.get_sheet_by_name(trans_sheets[0])
				book_trans.active

				# Adding Title for the columns
				main_sheet['D' + str(1)] = "New Version"
				main_sheet['E' + str(1)] = "Old Version"
				book_main.save(file_path)
				
				chapters(main_sheet, trans_sheet, book_main, file_path)
		except:
			print(m_f_name)