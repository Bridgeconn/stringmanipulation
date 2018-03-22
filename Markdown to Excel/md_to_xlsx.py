# Convert folders (book,chapters) and files from a main folder
# to an excel file. 


import openpyxl
import os
import glob
import re
#import pandas as pd
#import xlsxwriter
from openpyxl import load_workbook

def books(file_name):
	# Loads the workbook
	work_book = load_workbook(file_name)
	worksheet = work_book.get_sheet_by_name("Sheet")	
	# Updating the sheet with column
	worksheet['A1'] = 'Book'
	worksheet['B1'] = 'Chapter'
	worksheet['C1'] = 'Verse'
	worksheet['D1'] = 'Notes'
	work_book.save(file_name)
	
def book_update(rows,book_name,chapter,verse,notes,file_name):
	work_book = load_workbook(file_name)
	worksheet = work_book.get_sheet_by_name('Sheet')
	work_book.active
	# Writing the values row by row
	worksheet['A' + str(rows)] = book_name
	worksheet['A' + str(rows)].value
	worksheet['B' + str(rows)] = chapter
	worksheet['B' + str(rows)].value
	worksheet['C' + str(rows)] = verse
	worksheet['C' + str(rows)].value
	worksheet['D' + str(rows)] = notes
	worksheet['D' + str(rows)].value
	work_book.save(file_name)
	

# Accessing current directory 
cwd = os.getcwd()

# Fetching all folders
folder = glob.glob("**/")	

for book_folder in folder:
	# For writing in excel file, static row number
	rows = 2		
	book_name = re.sub(r"/","",book_folder)
	os.chdir(book_folder)

	# Fetching the folders of chapters
	book_folder = glob.glob("**/")	
	sort_folder = sorted(book_folder)

	for chapter_folder in sort_folder:
		chapter_number = re.sub(r"/","",chapter_folder)
		# Accessinf the directory
		os.chdir(chapter_folder)

		# Fetching .md files	
		chapter_folder = glob.glob("*.md")	
		sort_verse = sorted(chapter_folder)
		current_position = 0
		next_position = 0
	
		for verse_number in sort_verse:

			# To get the number of next verse 
			value = sort_verse[current_position % (len(sort_verse))]
			next_position += 1
			next_value = sort_verse[next_position % (len(sort_verse))]

			if value != next_value :
				next_verse = next_value
				x = 1
			else:
				x = 0

			verse_till = re.sub(r".md","",next_verse)
			verse_end = int(verse_till)
			versus = str(verse_end - 1)

			# "doc" reads the data from file
			doc = open(verse_number).read()

			# Trims the extension
			first_verse = re.sub(r".md","",verse_number)		
			replace_obj = re.sub(r' \n+ ',' - ',doc)	
			searchobj = re.sub(r'-?#','\n•',replace_obj)

			# Checks the digit of verse
			if len(first_verse) < 2:
				first_verse = "0" + first_verse

			if len(versus) < 2:
				versus = "0" + versus

			# "book1" is path of excel file where we store the content	
			book1 = cwd + "/" + book_name + ".xlsx"
			
			# Condition is used to mention verse to verse 
			if x == 1:
				verses = first_verse + " - " + versus
			else:
				verses = first_verse

			if os.path.exists(book1):
				book_update(rows,book_name,chapter_number,verses,searchobj,book1)
				rows +=1
			else:
				work_book = openpyxl.Workbook()	# Open a new workbook
				work_book.save(book1)
				books(book1)
				book_update(rows,book_name,chapter_number,verses,searchobj,book1)
				rows +=1

		os.chdir("..")
	os.chdir("..")